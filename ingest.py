# ── IMPORTS ──────────────────────────────────────────────────────────────────

import chromadb, sys, pathlib, requests
# chromadb   — the vector database library that stores document embeddings
# sys        — gives access to command line arguments (sys.argv)
# pathlib    — modern Python way to handle file paths cross-platform
# requests   — HTTP library for fetching web pages via URL

from langchain_text_splitters import RecursiveCharacterTextSplitter
# Splits large documents into smaller chunks before storing in ChromaDB
# "Recursive" means it tries to split on paragraphs first, then sentences,
# then words — preserving meaning rather than cutting mid-sentence

from chromadb.utils.embedding_functions import OllamaEmbeddingFunction
# Tells ChromaDB to use Ollama (running locally) to convert text into vectors
# Instead of sending text to OpenAI for embedding, everything stays on your Mac

# ── EMBEDDING FUNCTION ────────────────────────────────────────────────────────

embedding_fn = OllamaEmbeddingFunction(
    model_name="nomic-embed-text",  # The specific embedding model — converts text to 768-dimensional vectors
    url="http://localhost:11434/api/embeddings"  # Ollama's local API endpoint — never leaves your machine
)
# This object is passed to every ChromaDB collection so all vectors are
# created using the same model. Critically: ingest and query MUST use
# identical embedding functions or search results will be meaningless

# ── CHROMADB CLIENT ───────────────────────────────────────────────────────────

client = chromadb.PersistentClient(path="./vectorstore")
# Creates (or connects to) a ChromaDB database stored in the ./vectorstore folder
# PersistentClient means data survives restarts — it's written to disk not RAM

# ── TEXT SPLITTER ─────────────────────────────────────────────────────────────

splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
# chunk_size=500  — each chunk is maximum 500 characters
# chunk_overlap=50 — consecutive chunks share 50 characters
# Overlap prevents context being lost at chunk boundaries
# e.g. if a sentence starts at char 490 and ends at 530, both chunks contain it

# ── TEXT EXTRACTION ───────────────────────────────────────────────────────────

def extract_text(path: pathlib.Path) -> str:
    # Takes a file path, returns its text content as a plain string
    # Each file type needs a different extraction method — this function handles all of them
    
    ext = path.suffix.lower()
    # Gets the file extension e.g. ".pdf", ".xlsx", ".docx"
    # .lower() ensures ".PDF" and ".pdf" are treated identically
    
    if ext == ".pdf":
        import pdfplumber
        # pdfplumber reads PDF files — imported here (not at top) to avoid
        # loading it unless we actually need it (lazy import pattern)
        with pdfplumber.open(path) as pdf:
            return " ".join([p.extract_text() or "" for p in pdf.pages])
            # Loops through every page, extracts text, joins with spaces
            # "or ''" handles pages with no extractable text (e.g. scanned images)
            # Returns one long string of the entire PDF's text content

    elif ext in [".xlsx", ".xls"]:
        import openpyxl
        # openpyxl reads Excel files
        wb = openpyxl.load_workbook(path, data_only=True)
        # data_only=True returns cell values not formulas
        # e.g. returns 42 instead of =SUM(A1:A10)
        rows = []
        for sheet in wb.worksheets:
            # Loops through every tab/sheet in the workbook
            for row in sheet.iter_rows(values_only=True):
                # iter_rows returns each row as a tuple of cell values
                rows.append(" | ".join([str(c) for c in row if c is not None]))
                # Joins non-empty cells with " | " separator
                # str(c) converts numbers and dates to strings
                # if c is not None skips empty cells
        return "\n".join(rows)
        # Each row becomes a line — preserves the table structure as readable text

    elif ext == ".docx":
        from docx import Document
        # python-docx library reads Word documents
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
        # Each paragraph in the Word doc becomes a line
        # p.text extracts the raw text ignoring formatting

    elif ext in [".html", ".htm"]:
        from bs4 import BeautifulSoup
        # BeautifulSoup parses HTML and strips tags
        soup = BeautifulSoup(path.read_text(), "html.parser")
        return soup.get_text(separator=" ")
        # get_text() removes all HTML tags and returns clean readable text
        # separator=" " puts spaces between elements instead of running them together

    elif ext in [".txt", ".md", ".py", ".js", ".csv", ".json"]:
        return path.read_text(errors="ignore")
        # Plain text files — read directly with no special processing
        # errors="ignore" skips unreadable characters (e.g. unusual encoding)
        # Covers: text files, markdown, Python code, JavaScript, CSV, JSON

    else:
        return f"Unsupported file type: {ext}"
        # Returns an error string for unknown file types
        # ingest_file() checks for this string and aborts rather than storing garbage

# ── URL INGESTION ─────────────────────────────────────────────────────────────

def ingest_url(url: str):
    """Fetch a web page and ingest its text content — goes into work_docs by default."""
    from bs4 import BeautifulSoup
    
    resp = requests.get(url, timeout=10)
    # Fetches the web page — timeout=10 means give up after 10 seconds
    # Prevents the script hanging indefinitely on slow or dead URLs
    
    soup = BeautifulSoup(resp.text, "html.parser")
    # Parses the HTML response
    
    text = soup.get_text(separator=" ")
    # Strips all HTML tags — extracts only the readable text content
    
    name = url.split("/")[2]
    # Extracts the domain name from the URL for use as an identifier
    # e.g. "https://www.bbc.com/news" → "www.bbc.com"
    # Used to create unique chunk IDs so the same URL isn't duplicated
    
    chunks = splitter.split_text(text)
    # Splits the page text into 500-character chunks with 50-char overlap
    
    url_collection = client.get_or_create_collection(
        name="work_docs",           # Web pages go into the work domain by default
        embedding_function=embedding_fn  # Same embedding function used everywhere
    )
    
    url_collection.add(
        documents=chunks,           # The actual text chunks to store
        ids=[f"url_{name}_{i}" for i in range(len(chunks))],
        # Unique ID for each chunk: "url_www.bbc.com_0", "url_www.bbc.com_1" etc.
        # ChromaDB requires unique IDs — running the same URL twice would error without --delete first
        metadatas=[{"source": url, "chunk": i, "domain": "work_docs"}
                   for i in range(len(chunks))]
        # Metadata stored alongside each chunk
        # "source" lets you filter or delete by URL later
        # "chunk" is the position number within the document
        # "domain" identifies which collection it belongs to
    )
    print(f"✓ Ingested {len(chunks)} chunks from {url} → work_docs")

# ── FILE INGESTION ────────────────────────────────────────────────────────────

def ingest_file(filepath: str):
    path = pathlib.Path(filepath)
    # Converts string filepath to a Path object for reliable cross-platform handling
    
    if not path.exists():
        print(f"File not found: {filepath}"); return
        # Exits early if the file doesn't exist — prevents cryptic errors later

    parent = path.parent.name.lower()
    # Gets the name of the folder containing the file
    # e.g. documents/finance/report.pdf → parent = "finance"
    # .lower() makes it case-insensitive

    domain_map = {
        "personal": "personal_docs",       # documents/personal/ → personal_docs collection
        "finance":  "finance_docs",         # documents/finance/ → finance_docs collection
        "work":     "work_docs",            # documents/work/ → work_docs collection
        "financial_reports": "finance_docs" # documents/finance/Financial_Reports/ → also finance_docs
    }
    collection_name = domain_map.get(parent, "work_docs")
    # Looks up the folder name in the map
    # If the folder name isn't recognised, defaults to work_docs
    # This is the automatic domain detection — folder structure determines collection

    domain_collection = client.get_or_create_collection(
        name=collection_name,           # Creates the collection if it doesn't exist yet
        embedding_function=embedding_fn # Same embedding function — critical for consistency
    )

    text = extract_text(path)
    # Calls the extract_text function above to get raw text from the file
    
    if not text or text.startswith("Unsupported"):
        print(f"✗ Could not extract text from {path.name}: {text}")
        return
        # Aborts if extraction failed or returned an unsupported type error
        # Without this check, empty chunks would be stored silently

    chunks = splitter.split_text(text)
    # Splits the extracted text into 500-char chunks with 50-char overlap
    
    if not chunks:
        print(f"✗ No chunks created from {path.name} — file may be empty")
        return
        # Guards against empty files that extract to an empty string
        # Calling .add() with empty chunks would cause a ChromaDB error

    domain_collection.add(
        documents=chunks,
        ids=[f"{path.stem}_{i}" for i in range(len(chunks))],
        # path.stem is the filename without extension
        # e.g. "Anuj_Upadhyay.pdf" → "Anuj_Upadhyay_0", "Anuj_Upadhyay_1" etc.
        metadatas=[{"source": path.name, "chunk": i, "domain": collection_name}
                   for i in range(len(chunks))]
        # Stores the original filename so you can delete by filename later
    )
    print(f"✓ Ingested {len(chunks)} chunks from {path.name} → {collection_name}")

# ── DELETE SOURCE ─────────────────────────────────────────────────────────────

def delete_source(source_name: str, domain: str = "work_docs"):
    """Remove all chunks for a file from a specific domain collection."""
    # Used before re-ingesting an updated document
    # Without deleting first, you'd have duplicate chunks causing noisy search results
    
    domain_collection = client.get_or_create_collection(
        name=domain,
        embedding_function=embedding_fn
    )
    
    existing = domain_collection.get(where={"source": source_name})
    # Finds all chunks where the "source" metadata field matches the filename
    # This is why we store "source" in metadata during ingestion
    
    if existing["ids"]:
        domain_collection.delete(ids=existing["ids"])
        # Deletes all matching chunks by their unique IDs
        print(f"✓ Deleted {len(existing['ids'])} old chunks for {source_name} from {domain}")
    else:
        print(f"No existing chunks found for {source_name} in {domain}")
        # Useful feedback — tells you if the filename doesn't match what was ingested

# ── ENTRY POINT ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # This block only runs when you call python3 ingest.py directly
    # It does NOT run when another file imports ingest.py
    
    if len(sys.argv) < 2:
        # sys.argv is the list of command line arguments
        # sys.argv[0] is always the script name itself
        # Less than 2 means no arguments were provided
        print("Usage: python3 ingest.py documents/personal/file.pdf")
        print("       python3 ingest.py https://example.com")
        print("       python3 ingest.py --delete filename.pdf finance_docs")
    
    elif sys.argv[1] == "--delete":
        # First argument is "--delete" — user wants to remove a document
        if len(sys.argv) < 4:
            # Need at least 4 args: script, --delete, filename, domain
            print("Usage: python3 ingest.py --delete filename.pdf domain_name")
        else:
            delete_source(sys.argv[2], sys.argv[3])
            # sys.argv[2] = filename e.g. "Anuj_Upadhyay.pdf"
            # sys.argv[3] = domain e.g. "personal_docs"
    
    elif sys.argv[1].startswith("http"):
        # If the argument starts with "http" it's a URL not a file path
        ingest_url(sys.argv[1])
    
    else:
        # Everything else is treated as a local file path
        ingest_file(sys.argv[1])